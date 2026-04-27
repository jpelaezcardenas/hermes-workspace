#!/usr/bin/env python3
"""Shared validators for SettleMint office deliverables.

These checks are deliberately fail-closed.
- DOCX outputs must retain the expected template signature for critical XML parts.
- Bid-manager questionnaire XLSX outputs must use the approved locked workbook template.
"""

from __future__ import annotations

import csv
import hashlib
import zipfile
from pathlib import Path

from openpyxl import load_workbook


class OfficeValidationError(RuntimeError):
    """Raised when an office deliverable does not match the required guardrails."""


DOCX_SIGNATURE_PARTS = (
    "word/styles.xml",
    "word/numbering.xml",
    "word/settings.xml",
    "word/theme/theme1.xml",
)

APPROVED_QUESTIONNAIRE_TEMPLATE_SHA256 = "40c7b082182bb30074f78cd08fff2a0ae603e4f90b982442b37aed7ecbccfdca"
QUESTIONNAIRE_RESPONSE_SHEET = "Response"
QUESTIONNAIRE_RESPONSE_TITLE_CELL = "A1"
QUESTIONNAIRE_RESPONSE_HEADER_ROW = 2
QUESTIONNAIRE_RESPONSE_DATA_START_ROW = 3
QUESTIONNAIRE_STATUS_COLUMN = "F"
QUESTIONNAIRE_STATUS_VALIDATION_RANGE = "F3:F500"


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _file_sha256(path: Path) -> str:
    hasher = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            hasher.update(chunk)
    return hasher.hexdigest()


def _zip_part_hashes(path: Path, parts: tuple[str, ...]) -> dict[str, str]:
    if not path.exists():
        raise OfficeValidationError(f"Required file is missing: {path}")
    with zipfile.ZipFile(path) as zf:
        hashes: dict[str, str] = {}
        for part in parts:
            try:
                hashes[part] = _sha256(zf.read(part))
            except KeyError as exc:
                raise OfficeValidationError(f"{path.name} is missing required part: {part}") from exc
        return hashes


def _zip_part_hash(path: Path, part: str) -> str:
    return _zip_part_hashes(path, (part,))[part]


def _normalized_row(values: list[object]) -> list[str]:
    normalized: list[str] = []
    for value in values:
        if value is None:
            normalized.append("")
        else:
            normalized.append(str(value).strip())
    return normalized


def _sheet_nonempty_cells(ws) -> dict[str, str]:
    cells: dict[str, str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value)
            if not text.strip():
                continue
            cells[cell.coordinate] = text
    return cells


def validate_docx_against_template(
    output_path: str | Path,
    template_path: str | Path,
    *,
    allowed_signature_overrides: dict[str, set[str]] | None = None,
    required_document_tokens: tuple[str, ...] = (),
) -> Path:
    """Fail unless the generated DOCX preserves the required template signature."""
    output = Path(output_path)
    template = Path(template_path)

    if not template.exists():
        raise OfficeValidationError(f"Locked DOCX template is missing: {template}")
    if not output.exists():
        raise OfficeValidationError(f"Expected DOCX output was not created: {output}")

    template_hashes = _zip_part_hashes(template, DOCX_SIGNATURE_PARTS)
    output_hashes = _zip_part_hashes(output, DOCX_SIGNATURE_PARTS)

    mismatches = []
    for part in DOCX_SIGNATURE_PARTS:
        allowed_hashes = {template_hashes[part]}
        if allowed_signature_overrides and part in allowed_signature_overrides:
            allowed_hashes |= set(allowed_signature_overrides[part])
        if output_hashes[part] not in allowed_hashes:
            mismatches.append(
                f"{part}: allowed {sorted(allowed_hashes)}, got {output_hashes[part]}"
            )

    if required_document_tokens:
        with zipfile.ZipFile(output) as zf:
            document_xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
        missing_tokens = [token for token in required_document_tokens if token not in document_xml]
        if missing_tokens:
            mismatches.append(f"word/document.xml is missing required template markers: {missing_tokens}")

    if mismatches:
        mismatch_text = "\n  - ".join([""] + mismatches)
        raise OfficeValidationError(
            "Generated DOCX failed template validation. "
            "This usually means the wrong template or a non-canonical fallback was used."
            f"{mismatch_text}"
        )

    return output


def read_csv_rows(path: str | Path) -> list[list[str]]:
    src = Path(path)
    if not src.exists():
        raise OfficeValidationError(f"Input CSV is missing: {src}")
    with src.open(encoding="utf-8", newline="") as handle:
        return list(csv.reader(handle))


def validate_csv_has_header(rows: list[list[str]], source: str | Path) -> None:
    if not rows:
        raise OfficeValidationError(f"CSV input is empty: {source}")
    header = rows[0]
    if not header or not any((cell or "").strip() for cell in header):
        raise OfficeValidationError(f"CSV input has no usable header row: {source}")


def validate_questionnaire_template(template_path: str | Path) -> Path:
    template = Path(template_path)
    if not template.exists():
        raise OfficeValidationError(f"Locked questionnaire template is missing: {template}")

    template_hash = _file_sha256(template)
    if template_hash != APPROVED_QUESTIONNAIRE_TEMPLATE_SHA256:
        raise OfficeValidationError(
            "Locked questionnaire template does not match the approved workbook. "
            f"Expected {APPROVED_QUESTIONNAIRE_TEMPLATE_SHA256}, got {template_hash}"
        )

    wb = load_workbook(template)
    expected_sheets = ["Cover", "Response", "Evidence Index", "Glossary"]
    if wb.sheetnames != expected_sheets:
        raise OfficeValidationError(
            "Locked questionnaire template has the wrong sheet contract: "
            f"expected {expected_sheets}, got {wb.sheetnames}"
        )

    response_ws = wb[QUESTIONNAIRE_RESPONSE_SHEET]
    if response_ws.freeze_panes != "A3":
        raise OfficeValidationError(
            f"Locked questionnaire template must freeze the response sheet at A3, got {response_ws.freeze_panes}"
        )

    return template


def validate_xlsx_expected_shape(
    output_path: str | Path,
    *,
    template_path: str | Path,
    expected_response_headers: list[str],
    expected_response_rows: int,
) -> Path:
    """Fail unless the generated workbook preserves the approved questionnaire contract."""
    output = Path(output_path)
    template = validate_questionnaire_template(template_path)

    if not output.exists():
        raise OfficeValidationError(f"Expected XLSX output was not created: {output}")

    mismatches: list[str] = []

    template_wb = load_workbook(template)
    output_wb = load_workbook(output)

    if output_wb.sheetnames != template_wb.sheetnames:
        mismatches.append(
            f"Sheet names/order must match template exactly: expected {template_wb.sheetnames}, got {output_wb.sheetnames}"
        )

    for sheet_name in template_wb.sheetnames:
        template_ws = template_wb[sheet_name]
        if sheet_name not in output_wb.sheetnames:
            mismatches.append(f"Missing required sheet '{sheet_name}'")
            continue
        output_ws = output_wb[sheet_name]

        if output_ws.freeze_panes != template_ws.freeze_panes:
            mismatches.append(
                f"Sheet '{sheet_name}' freeze panes mismatch: expected {template_ws.freeze_panes}, got {output_ws.freeze_panes}"
            )
        if output_ws.auto_filter.ref != template_ws.auto_filter.ref:
            mismatches.append(
                f"Sheet '{sheet_name}' filter range mismatch: expected {template_ws.auto_filter.ref}, got {output_ws.auto_filter.ref}"
            )
        if sorted(str(r) for r in output_ws.merged_cells.ranges) != sorted(str(r) for r in template_ws.merged_cells.ranges):
            mismatches.append(f"Sheet '{sheet_name}' merged-cell contract changed")
        if output_ws.max_row != template_ws.max_row or output_ws.max_column != template_ws.max_column:
            mismatches.append(
                f"Sheet '{sheet_name}' dimensions changed: expected {template_ws.max_row}x{template_ws.max_column}, "
                f"got {output_ws.max_row}x{output_ws.max_column}"
            )

    response_template_ws = template_wb[QUESTIONNAIRE_RESPONSE_SHEET]
    if QUESTIONNAIRE_RESPONSE_SHEET not in output_wb.sheetnames:
        mismatches.append(f"Missing required sheet '{QUESTIONNAIRE_RESPONSE_SHEET}'")
        response_output_ws = None
    else:
        response_output_ws = output_wb[QUESTIONNAIRE_RESPONSE_SHEET]
    data_capacity = response_template_ws.max_row - QUESTIONNAIRE_RESPONSE_DATA_START_ROW + 1
    if expected_response_rows > data_capacity:
        mismatches.append(
            f"Response row count {expected_response_rows} exceeds template capacity {data_capacity}"
        )

    expected_headers = _normalized_row(expected_response_headers)
    template_headers = _normalized_row(
        [response_template_ws.cell(row=QUESTIONNAIRE_RESPONSE_HEADER_ROW, column=col).value for col in range(1, response_template_ws.max_column + 1)]
    )

    if response_output_ws is not None:
        title_cell = QUESTIONNAIRE_RESPONSE_TITLE_CELL
        if response_output_ws[title_cell].value != response_template_ws[title_cell].value:
            mismatches.append(
                f"Response title cell {title_cell} must remain '{response_template_ws[title_cell].value}'"
            )

        actual_headers = _normalized_row(
            [response_output_ws.cell(row=QUESTIONNAIRE_RESPONSE_HEADER_ROW, column=col).value for col in range(1, response_output_ws.max_column + 1)]
        )
        if actual_headers != expected_headers:
            mismatches.append(f"Response headers/order mismatch: expected {expected_headers}, got {actual_headers}")
        if actual_headers != template_headers:
            mismatches.append(f"Response headers must match template exactly: expected {template_headers}, got {actual_headers}")

        for col in range(1, response_output_ws.max_column + 1):
            output_cell = response_output_ws.cell(row=QUESTIONNAIRE_RESPONSE_HEADER_ROW, column=col)
            template_cell = response_template_ws.cell(row=QUESTIONNAIRE_RESPONSE_HEADER_ROW, column=col)
            if output_cell.style_id != template_cell.style_id:
                mismatches.append(f"Response header style changed at {output_cell.coordinate}")

        first_blank_row = QUESTIONNAIRE_RESPONSE_DATA_START_ROW + expected_response_rows
        for row in range(first_blank_row, response_output_ws.max_row + 1):
            if any(
                response_output_ws.cell(row=row, column=col).value not in (None, "")
                for col in range(1, response_output_ws.max_column + 1)
            ):
                mismatches.append(
                    f"Response row {row} must be blank after the populated data range"
                )
                break

        data_validations = response_output_ws.data_validations.dataValidation
        if not data_validations:
            mismatches.append("Response sheet is missing status data validation")
        else:
            if not any(str(dv.sqref) == QUESTIONNAIRE_STATUS_VALIDATION_RANGE for dv in data_validations):
                mismatches.append(
                    f"Response status validation range must remain {QUESTIONNAIRE_STATUS_VALIDATION_RANGE}"
                )

    for sheet_name in ("Cover", "Evidence Index", "Glossary"):
        if sheet_name not in output_wb.sheetnames:
            continue
        if _sheet_nonempty_cells(output_wb[sheet_name]) != _sheet_nonempty_cells(template_wb[sheet_name]):
            mismatches.append(f"Locked sheet '{sheet_name}' content changed unexpectedly")

    if mismatches:
        mismatch_text = "\n  - ".join([""] + mismatches)
        raise OfficeValidationError(
            "Generated XLSX failed the locked questionnaire template contract. "
            "This usually means a non-template writer path or a structural workbook mutation was used."
            f"{mismatch_text}"
        )

    return output
