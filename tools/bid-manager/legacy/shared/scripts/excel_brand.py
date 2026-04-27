#!/usr/bin/env python3
"""Shared Excel branding utilities for SettleMint office agents.

Provides consistent Figtree-based, lavender-accented styling that matches
the DOCX brand system (markdown_to_docx.py color palette).

Usage:
    from excel_brand import format_entire_sheet, create_branded_workbook
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PrintPageSetup

# ---------------------------------------------------------------------------
# Brand colors (same as DOCX markdown_to_docx.py)
# ---------------------------------------------------------------------------
NAVY = "000099"           # Heading text color
DARK_TEXT = "1F1F1F"      # Body text color
HEADER_FILL = "E8E0F0"   # Light lavender for header rows
HEADER_TEXT = "2D1A8C"    # Dark purple/indigo for header text
BORDER_COLOR = "C4B4D8"  # Medium lavender borders
ALT_ROW_FILL = "F8F6FC"  # Very subtle lavender for alternating rows
WHITE = "FFFFFF"

# ---------------------------------------------------------------------------
# Font settings
# ---------------------------------------------------------------------------
FONT_NAME = "Figtree"
HEADER_FONT_SIZE = 12
BODY_FONT_SIZE = 11
TITLE_FONT_SIZE = 20
SUBTITLE_FONT_SIZE = 14

# ---------------------------------------------------------------------------
# Logo path
# ---------------------------------------------------------------------------
LOGO_PATH = Path(__file__).resolve().parents[1] / "brand" / "settlemint-logo-dark.png"


# ===================================================================
# Primitive builders
# ===================================================================

def get_brand_font(size: int = 11, bold: bool = False, color: str = DARK_TEXT) -> Font:
    """Return an openpyxl Font with Figtree and the given attributes."""
    return Font(name=FONT_NAME, size=size, bold=bold, color=color)


def get_header_fill() -> PatternFill:
    """Return lavender header fill."""
    return PatternFill(start_color=HEADER_FILL, end_color=HEADER_FILL, fill_type="solid")


def _alt_row_fill() -> PatternFill:
    return PatternFill(start_color=ALT_ROW_FILL, end_color=ALT_ROW_FILL, fill_type="solid")


def _white_fill() -> PatternFill:
    return PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")


def get_thin_border() -> Border:
    """Return thin lavender border on all four sides."""
    side = Side(style="thin", color=BORDER_COLOR)
    return Border(left=side, right=side, top=side, bottom=side)


# ===================================================================
# Row / sheet formatters
# ===================================================================

def format_header_row(ws, row_num: int, col_start: int = 1, col_end: int | None = None) -> None:
    """Apply branded header styling to *row_num*.

    Lavender fill, Figtree 12 pt bold dark-purple text, thin borders,
    center alignment (first column left-aligned), and auto-filter.
    """
    if col_end is None:
        col_end = ws.max_column or 1

    header_font = get_brand_font(size=HEADER_FONT_SIZE, bold=True, color=HEADER_TEXT)
    fill = get_header_fill()
    border = get_thin_border()

    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = fill
        cell.border = border
        if col == col_start:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Auto-filter spanning the header range
    start_letter = get_column_letter(col_start)
    end_letter = get_column_letter(col_end)
    ws.auto_filter.ref = f"{start_letter}{row_num}:{end_letter}{ws.max_row}"


def format_data_rows(
    ws,
    start_row: int,
    end_row: int | None = None,
    col_start: int = 1,
    col_end: int | None = None,
) -> None:
    """Apply branded styling to data rows.

    Figtree 11 pt, dark text, text wrapping, top alignment, thin borders,
    alternating row fill (subtle lavender on even rows, white on odd).
    """
    if end_row is None:
        end_row = ws.max_row or start_row
    if col_end is None:
        col_end = ws.max_column or 1

    body_font = get_brand_font(size=BODY_FONT_SIZE)
    border = get_thin_border()
    alt_fill = _alt_row_fill()
    white_fill = _white_fill()

    for row in range(start_row, end_row + 1):
        fill = alt_fill if (row - start_row) % 2 == 1 else white_fill
        for col in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def auto_column_widths(
    ws,
    min_width: int = 10,
    max_width: int = 50,
    col_start: int = 1,
    col_end: int | None = None,
) -> None:
    """Auto-size columns based on content length (scans first 50 rows)."""
    if col_end is None:
        col_end = ws.max_column or 1

    scan_rows = min(ws.max_row or 1, 50)

    for col in range(col_start, col_end + 1):
        best = min_width
        for row in range(1, scan_rows + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                # Use the longest line if multi-line
                for line in str(val).split("\n"):
                    best = max(best, len(line) + 2)
        ws.column_dimensions[get_column_letter(col)].width = min(best, max_width)


# ===================================================================
# Logo helper
# ===================================================================

def add_logo(ws, cell: str = "A1", width_px: int = 200) -> None:
    """Insert the SettleMint logo anchored to *cell*."""
    if not LOGO_PATH.exists():
        return  # Graceful skip if logo missing
    img = XlImage(str(LOGO_PATH))
    # Scale proportionally
    aspect = img.height / img.width if img.width else 1
    img.width = width_px
    img.height = int(width_px * aspect)
    ws.add_image(img, cell)


# ===================================================================
# Workbook-level helpers
# ===================================================================

def create_branded_workbook(
    title: str,
    subtitle: str | None = None,
    sheets: list[str] | None = None,
) -> Workbook:
    """Create a new branded workbook with a cover sheet and optional data sheets.

    Returns an unsaved Workbook ready for population.
    """
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"

    # Logo
    add_logo(cover, "A1", width_px=200)

    # Title in B3
    title_cell = cover.cell(row=3, column=2, value=title)
    title_cell.font = get_brand_font(size=TITLE_FONT_SIZE, bold=True, color=NAVY)
    title_cell.alignment = Alignment(vertical="center")

    # Subtitle in B4
    if subtitle:
        sub_cell = cover.cell(row=4, column=2, value=subtitle)
        sub_cell.font = get_brand_font(size=SUBTITLE_FONT_SIZE, color=NAVY)
        sub_cell.alignment = Alignment(vertical="center")

    # Widen column B for title text
    cover.column_dimensions["B"].width = 60

    # Additional sheets
    if sheets:
        for name in sheets:
            ws = wb.create_sheet(title=name)
            # Default print setup
            _apply_print_setup(ws)

    return wb


def _apply_print_setup(ws) -> None:
    """Configure landscape, fit-to-width, margins."""
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75


# ===================================================================
# Convenience composites
# ===================================================================

def format_entire_sheet(ws, header_row: int = 1, data_start_row: int = 2) -> None:
    """One-call convenience: header + data rows + column widths + freeze panes + print setup."""
    col_end = ws.max_column or 1
    data_end_row = ws.max_row or data_start_row

    format_header_row(ws, header_row, col_end=col_end)
    format_data_rows(ws, data_start_row, end_row=data_end_row, col_end=col_end)
    auto_column_widths(ws, col_end=col_end)

    # Freeze panes just below the header row
    ws.freeze_panes = ws.cell(row=data_start_row, column=1)

    # Print: repeat header row on every page
    ws.print_title_rows = f"{header_row}:{header_row}"

    _apply_print_setup(ws)


def apply_brand_to_existing(
    ws,
    header_row: int,
    data_start_row: int,
    data_end_row: int | None = None,
) -> None:
    """Brand an existing/template-cloned worksheet without overwriting header formatting.

    Applies Figtree font, alignment, borders, and alternating fill to data cells only.
    """
    if data_end_row is None:
        data_end_row = ws.max_row or data_start_row
    col_end = ws.max_column or 1

    format_data_rows(ws, data_start_row, end_row=data_end_row, col_end=col_end)

    # Freeze at data start row
    ws.freeze_panes = ws.cell(row=data_start_row, column=1)

    # Print: repeat header row
    ws.print_title_rows = f"{header_row}:{header_row}"

    _apply_print_setup(ws)
