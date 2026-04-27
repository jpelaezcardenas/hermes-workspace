from __future__ import annotations

from pathlib import Path
import shutil

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.formatting.rule import Rule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation


BASE_DIR = Path("/Users/quark/Public/quark/workspace")
OUTPUT_DIR = BASE_DIR / "bid-manager/skeletons/5_questionnaire/excel"
VERSION = "v1.0.0"
VERSIONED_OUTPUT_PATH = OUTPUT_DIR / f"settlemint-questionnaire-template-{VERSION}.xlsx"
OUTPUT_PATH = OUTPUT_DIR / "settlemint-questionnaire-template.xlsx"
SCRIPT_DIR = BASE_DIR / "bid-manager/scripts"

NAVY = "000099"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
LIGHT_BLUE = "D6E4F0"
ACCENT = "4472C4"
GREEN_FILL = "C6EFCE"
GREEN_TEXT = "006100"
YELLOW_FILL = "FFEB9C"
YELLOW_TEXT = "9C5700"
BLUE_FILL = "BDD7EE"
BLUE_TEXT = "1F4E79"
RED_FILL = "FFC7CE"
RED_TEXT = "9C0006"
NA_FILL = "D9D9D9"
NA_TEXT = "595959"
BORDER_COLOR = "C0C0C0"
TEXT_DARK = "1F1F1F"
TEXT_MUTED = "666666"

STATUS_OPTIONS = ["Compliant", "Partial", "Roadmap", "Non-Compliant", "N/A"]
MAX_DATA_ROW = 500
STATUS_COLUMN = "F"
DATA_START_ROW = 3


THIN_SIDE = Side(style="thin", color=BORDER_COLOR)
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
NO_BORDER = Border()
ALT_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
TITLE_FILL = PatternFill("solid", fgColor=LIGHT_BLUE)
WHITE_FILL = PatternFill("solid", fgColor=WHITE)

BODY_FONT = Font(name="Figtree", size=11, color=TEXT_DARK)
TITLE_BAR_FONT = Font(name="Figtree", size=14, bold=True, color=NAVY)
COLUMN_HEADER_FONT = Font(name="Figtree", size=12, bold=True, color=NAVY)
COVER_TITLE_FONT = Font(name="Figtree", size=20, bold=True, color=NAVY)
COVER_SUBTITLE_FONT = Font(name="Figtree", size=13, color=ACCENT)
SECTION_TITLE_FONT = Font(name="Figtree", size=16, bold=True, color=NAVY)
LABEL_FONT = Font(name="Figtree", size=11, bold=True, color=NAVY)
DISCLAIMER_FONT = Font(name="Figtree", size=11, italic=True, color=TEXT_MUTED)
PLACEHOLDER_FONT = Font(name="Figtree", size=11, italic=True, color="7F7F7F")
VERSION_FONT = Font(name="Figtree", size=9, italic=True, color="808080")

STATUS_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
TEXT_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
TEXT_MIDDLE_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)


def find_logo() -> Path | None:
    candidates = [
        BASE_DIR / "bid-manager/templates/settlemint-logo.png",
        BASE_DIR / "assets/logos/settlemint-logo.png",
        BASE_DIR / "assets/settlemint-logo-candidate.png",
    ]
    search_roots = [BASE_DIR / "bid-manager/templates", BASE_DIR / "assets"]

    for candidate in candidates:
        if candidate.exists():
            return candidate

    for root in search_roots:
        if not root.exists():
            continue
        for pattern in ("*settlemint*logo*.png", "*settlemint*.png", "*logo*.png"):
            matches = sorted(root.rglob(pattern))
            if matches:
                return matches[0]
    return None


def set_workbook_defaults(wb: Workbook) -> None:
    wb.properties.creator = "Quark"
    wb.properties.title = f"SettleMint Questionnaire Response Template {VERSION}"
    wb.properties.subject = "Questionnaire / RFI response workbook"
    wb.properties.company = "SettleMint NV"
    wb.properties.description = f"Branded Excel template for questionnaire and RFI responses ({VERSION})."


def style_cell(cell, *, font=None, fill=None, alignment=None, border=THIN_BORDER, number_format="General"):
    if font is not None:
        cell.font = font
    if fill is not None:
        cell.fill = fill
    if alignment is not None:
        cell.alignment = alignment
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def format_data_table(ws, start_row: int, end_row: int, end_col: int, status_col: int | None = None):
    for row in range(start_row, end_row + 1):
        fill = ALT_FILL if row % 2 == 1 else WHITE_FILL
        if ws.row_dimensions[row].height is None or ws.row_dimensions[row].height < 52:
            ws.row_dimensions[row].height = 52
        for col in range(1, end_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is None:
                cell.value = ""
            alignment = STATUS_ALIGNMENT if status_col and col == status_col else TEXT_ALIGNMENT
            style_cell(cell, font=BODY_FONT, fill=fill, alignment=alignment, border=THIN_BORDER)


def add_status_validation_and_formatting(ws, cell_range: str):
    dv = DataValidation(type="list", formula1='"' + ",".join(STATUS_OPTIONS) + '"', allow_blank=True)
    dv.prompt = "Select a compliance status"
    dv.error = "Choose one of: Compliant, Partial, Roadmap, Non-Compliant, N/A"
    ws.add_data_validation(dv)
    dv.add(cell_range)

    status_fill_fonts = [
        ("Compliant", GREEN_FILL, GREEN_TEXT),
        ("Partial", YELLOW_FILL, YELLOW_TEXT),
        ("Roadmap", BLUE_FILL, BLUE_TEXT),
        ("Non-Compliant", RED_FILL, RED_TEXT),
        ("N/A", NA_FILL, NA_TEXT),
    ]
    for text, fill, font_color in status_fill_fonts:
        dxf = DifferentialStyle(
            fill=PatternFill("solid", fgColor=fill),
            font=Font(name="Figtree", size=11, color=font_color, bold=True),
        )
        ws.conditional_formatting.add(
            cell_range,
            Rule(type="cellIs", operator="equal", formula=[f'"{text}"'], dxf=dxf),
        )


def create_cover_sheet(wb: Workbook, logo_path: Path | None):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A1"

    for col, width in {
        "A": 4,
        "B": 18,
        "C": 18,
        "D": 20,
        "E": 20,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 20,
        "J": 18,
        "K": 18,
        "L": 16,
    }.items():
        ws.column_dimensions[col].width = width

    for row, height in {
        1: 18,
        2: 18,
        3: 32,
        4: 32,
        5: 18,
        6: 26,
        7: 18,
        8: 28,
        9: 18,
        10: 24,
        11: 24,
        12: 24,
        13: 24,
        14: 24,
        15: 24,
        16: 18,
        17: 28,
        18: 28,
        19: 18,
        20: 18,
    }.items():
        ws.row_dimensions[row].height = height

    ws.merge_cells("B3:F4")
    banner = ws["B3"]
    banner.value = "SettleMint"
    style_cell(
        banner,
        font=COVER_TITLE_FONT,
        alignment=Alignment(vertical="center", horizontal="left", indent=1),
        fill=WHITE_FILL,
        border=NO_BORDER,
    )

    if logo_path and logo_path.exists():
        img = XLImage(str(logo_path))
        img.width = 350
        img.height = 80
        ws.add_image(img, "H3")
    else:
        ws.merge_cells("H3:L4")
        placeholder = ws["H3"]
        placeholder.value = "[SettleMint logo placeholder]"
        style_cell(
            placeholder,
            font=PLACEHOLDER_FONT,
            alignment=CENTER_ALIGNMENT,
            fill=WHITE_FILL,
            border=NO_BORDER,
        )

    ws.merge_cells("B6:L6")
    subtitle = ws["B6"]
    subtitle.value = "Digital Asset Lifecycle Platform (DALP)"
    style_cell(subtitle, font=COVER_SUBTITLE_FONT, alignment=TEXT_MIDDLE_ALIGNMENT, border=NO_BORDER, fill=WHITE_FILL)

    ws.merge_cells("B8:L8")
    title = ws["B8"]
    title.value = "Questionnaire Response"
    style_cell(title, font=SECTION_TITLE_FONT, alignment=TEXT_MIDDLE_ALIGNMENT, border=NO_BORDER, fill=WHITE_FILL)

    fields = [
        (10, "Client Name", "[VARIABLE: Client Name]", "G10", "[VARIABLE: Date]", "Date"),
        (11, "RFP/RFI Reference", "[VARIABLE: Reference Number]", "G11", "[VARIABLE: Version]", "Version"),
        (12, "Classification", "[VARIABLE: Confidential/Internal/Public]", "G12", "SettleMint NV", "Prepared by"),
        (13, "Contact", "[VARIABLE: Contact Name and Email]", None, None, None),
    ]

    for row, left_label, left_value, right_cell, right_value, right_label in fields:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=6)
        label_cell = ws.cell(row=row, column=2)
        value_cell = ws.cell(row=row, column=4)
        label_cell.value = left_label
        value_cell.value = left_value
        style_cell(label_cell, font=LABEL_FONT, fill=ALT_FILL, alignment=TEXT_MIDDLE_ALIGNMENT, border=THIN_BORDER)
        style_cell(value_cell, font=BODY_FONT, fill=WHITE_FILL, alignment=TEXT_MIDDLE_ALIGNMENT, border=THIN_BORDER)

        if right_cell:
            right_col = ws[right_cell].column
            ws.merge_cells(start_row=row, start_column=right_col, end_row=row, end_column=right_col + 1)
            ws.merge_cells(start_row=row, start_column=right_col + 2, end_row=row, end_column=right_col + 4)
            r_label = ws.cell(row=row, column=right_col)
            r_value = ws.cell(row=row, column=right_col + 2)
            r_label.value = right_label
            r_value.value = right_value
            style_cell(r_label, font=LABEL_FONT, fill=ALT_FILL, alignment=TEXT_MIDDLE_ALIGNMENT, border=THIN_BORDER)
            style_cell(r_value, font=BODY_FONT, fill=WHITE_FILL, alignment=TEXT_MIDDLE_ALIGNMENT, border=THIN_BORDER)

    ws.merge_cells("B14:C15")
    prepared_for_label = ws["B14"]
    prepared_for_label.value = "Prepared for"
    style_cell(prepared_for_label, font=LABEL_FONT, fill=ALT_FILL, alignment=TEXT_MIDDLE_ALIGNMENT, border=THIN_BORDER)

    ws.merge_cells("D14:L15")
    prepared_for_value = ws["D14"]
    prepared_for_value.value = "[VARIABLE: Client/Recipient Organisation]"
    style_cell(prepared_for_value, font=BODY_FONT, fill=WHITE_FILL, alignment=TEXT_MIDDLE_ALIGNMENT, border=THIN_BORDER)

    ws.merge_cells("B17:L18")
    disclaimer = ws["B17"]
    disclaimer.value = (
        "This document contains confidential and proprietary information of SettleMint NV. "
        "Distribution is restricted to the intended recipient."
    )
    style_cell(
        disclaimer,
        font=DISCLAIMER_FONT,
        alignment=TEXT_MIDDLE_ALIGNMENT,
        fill=ALT_FILL,
        border=THIN_BORDER,
    )

    ws.merge_cells("J20:L20")
    version_cell = ws["J20"]
    version_cell.value = f"Template {VERSION}"
    style_cell(
        version_cell,
        font=VERSION_FONT,
        alignment=Alignment(horizontal="right", vertical="center"),
        fill=WHITE_FILL,
        border=NO_BORDER,
    )

    ws.print_options.gridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.print_area = "A1:L20"


def create_response_sheet(wb: Workbook):
    ws = wb.create_sheet("Response")
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False
    ws.print_title_rows = "$1:$2"

    widths = {
        "A": 10,
        "B": 22,
        "C": 22,
        "D": 55,
        "E": 70,
        "F": 16,
        "G": 24,
        "H": 16,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A1:H1")
    head = ws["A1"]
    head.value = "SettleMint — Questionnaire Response"
    style_cell(head, font=TITLE_BAR_FONT, fill=TITLE_FILL, alignment=CENTER_ALIGNMENT, border=THIN_BORDER)

    headers = [
        "Ref #",
        "Category",
        "Sub-Category",
        "Question / Requirement",
        "SettleMint Response",
        "Status",
        "Evidence / Reference",
        "Notes",
    ]
    for idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=idx)
        cell.value = value
        style_cell(cell, font=COLUMN_HEADER_FONT, fill=TITLE_FILL, alignment=CENTER_ALIGNMENT, border=THIN_BORDER)

    sample_rows = [
        ["1.1", "Company", "Corporate", "Company incorporated in which jurisdiction?", "SettleMint NV is incorporated in Belgium and operates as a technology platform provider for digital asset infrastructure.", "Compliant", "EV-003", ""],
        ["2.1", "Security", "Encryption", "Describe your encryption standards for data at rest", "Data at rest protections are implemented using industry-standard encryption controls aligned to enterprise deployment requirements and hosting architecture.", "Compliant", "EV-001", "Tailor to environment-specific controls if needed."],
        ["3.1", "Compliance", "KYC/AML", "Do you support KYC/AML compliance checks?", "DALP supports compliance workflows and integrations that enable identity, screening, and policy-driven controls required for regulated digital asset operations.", "Compliant", "EV-003", "Reference OnchainID where relevant."],
        ["4.1", "Integration", "APIs", "List supported API protocols", "DALP exposes API-driven integration patterns for enterprise systems, enabling orchestration across lifecycle events, servicing, and ecosystem connectivity.", "Compliant", "EV-003", "Add client-specific interface details."],
        ["5.1", "Deployment", "Infrastructure", "Can the platform be deployed on-premises?", "DALP supports flexible deployment models, including customer-controlled environments, subject to target architecture and operational requirements.", "Compliant", "EV-003", "Clarify preferred deployment topology."],
        ["6.1", "Support", "SLA", "What are your SLA response times?", "Support coverage and response commitments are defined contractually based on service scope, environment criticality, and selected support tier.", "Partial", "EV-002", "Insert approved SLA wording before submission."],
    ]

    for r_idx, row in enumerate(sample_rows, start=DATA_START_ROW):
        ws.row_dimensions[r_idx].height = 52
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value

    format_data_table(ws, DATA_START_ROW, MAX_DATA_ROW, 8, status_col=6)
    add_status_validation_and_formatting(ws, f"{STATUS_COLUMN}{DATA_START_ROW}:{STATUS_COLUMN}{MAX_DATA_ROW}")

    ws.auto_filter.ref = f"A2:H{MAX_DATA_ROW}"
    ws.print_area = f"A1:H{MAX_DATA_ROW}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def create_evidence_sheet(wb: Workbook):
    ws = wb.create_sheet("Evidence Index")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.print_title_rows = "$1:$2"

    for col, width in {"A": 12, "B": 40, "C": 50, "D": 40}.items():
        ws.column_dimensions[col].width = width

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "Evidence Index"
    style_cell(title, font=TITLE_BAR_FONT, fill=TITLE_FILL, alignment=CENTER_ALIGNMENT, border=THIN_BORDER)

    headers = ["Evidence ID", "Document Title", "Description", "Location / URL"]
    for idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=idx)
        cell.value = value
        style_cell(cell, font=COLUMN_HEADER_FONT, fill=TITLE_FILL, alignment=CENTER_ALIGNMENT, border=THIN_BORDER)

    rows = [
        ["EV-001", "SettleMint Security Whitepaper", "Platform security architecture and controls", "[Link]"],
        ["EV-002", "SOC 2 Type II Report", "Independent audit report", "Available on request"],
        ["EV-003", "DALP Architecture Documentation", "Technical architecture overview", "[Link]"],
    ]
    for r_idx, row in enumerate(rows, start=3):
        ws.row_dimensions[r_idx].height = 52
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value

    format_data_table(ws, 3, 100, 4)
    ws.auto_filter.ref = "A2:D100"
    ws.print_area = "A1:D100"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def create_glossary_sheet(wb: Workbook):
    ws = wb.create_sheet("Glossary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"
    ws.print_title_rows = "$1:$2"

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 80
    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 28

    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "Glossary"
    style_cell(title, font=TITLE_BAR_FONT, fill=TITLE_FILL, alignment=CENTER_ALIGNMENT, border=THIN_BORDER)

    headers = ["Term", "Definition"]
    for idx, value in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=idx)
        cell.value = value
        style_cell(cell, font=COLUMN_HEADER_FONT, fill=TITLE_FILL, alignment=CENTER_ALIGNMENT, border=THIN_BORDER)

    rows = [
        ["DALP", "Digital Asset Lifecycle Platform — SettleMint's end-to-end platform for creating, managing, and servicing digital assets"],
        ["OnchainID", "On-chain identity framework enabling verifiable, claim-based identity for regulatory compliance"],
        ["EVM", "Ethereum Virtual Machine — the runtime environment for smart contracts on Ethereum-compatible networks"],
        ["DLT", "Distributed Ledger Technology — decentralized database technology underlying blockchain networks"],
        ["KYC", "Know Your Customer — identity verification process required by financial regulations"],
        ["AML", "Anti-Money Laundering — regulatory framework for preventing financial crime"],
    ]
    for r_idx, row in enumerate(rows, start=3):
        ws.row_dimensions[r_idx].height = 52
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx).value = value

    format_data_table(ws, 3, 100, 2)
    ws.auto_filter.ref = "A2:B100"
    ws.print_area = "A1:B100"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    SCRIPT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    set_workbook_defaults(wb)
    logo_path = find_logo()

    create_cover_sheet(wb, logo_path)
    create_response_sheet(wb)
    create_evidence_sheet(wb)
    create_glossary_sheet(wb)

    wb.save(VERSIONED_OUTPUT_PATH)
    shutil.copy2(VERSIONED_OUTPUT_PATH, OUTPUT_PATH)
    print(f"Generated: {VERSIONED_OUTPUT_PATH}")
    print(f"Copied to: {OUTPUT_PATH}")
    print(f"Logo used: {logo_path if logo_path else 'None (placeholder inserted)'}")


if __name__ == "__main__":
    main()
