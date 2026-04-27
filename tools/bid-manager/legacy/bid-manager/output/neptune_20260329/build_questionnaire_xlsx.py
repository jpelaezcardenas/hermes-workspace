#!/usr/bin/env python3
"""Build a polished Neptune questionnaire XLSX from the existing CSV."""

import csv
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT_CSV = "/Users/quark/Public/quark/workspace/quark-unsorted-output/Neptune_Questionnaire_Response_DALP.csv"
OUTPUT_XLSX = "/Users/quark/Public/quark/workspace/settlemint-office-agents/bid-manager/output/neptune_20260329/Neptune_Questionnaire_Response_DALP_v2.xlsx"

# Read CSV
rows = []
with open(INPUT_CSV, "r", encoding="utf-8") as f:
    reader = csv.reader(f)
    for row in reader:
        rows.append(row)

wb = Workbook()
ws = wb.active
ws.title = "Platform Functionalities"

# Styles
header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
cat_font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
cat_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
yes_font = Font(name="Calibri", bold=True, color="006100")
yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
no_font = Font(name="Calibri", bold=True, color="9C0006")
no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
partial_font = Font(name="Calibri", bold=True, color="9C6500")
partial_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
data_font = Font(name="Calibri", size=10)
wrap_align = Alignment(wrap_text=True, vertical="top")
thin_border = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

# Identify header row (row with "#", "Category", etc.)
header_idx = None
category_rows = set()
for i, row in enumerate(rows):
    if len(row) > 3 and row[1] == "#" and row[2] == "Category":
        header_idx = i
    # Category separator rows (all caps in col 1, no number)
    if len(row) > 1 and row[1] and row[1].isupper() and not row[1].isdigit() and row[1] != "#" and "&" in row[1] or (len(row) > 1 and row[1] and row[1].replace(" ", "").replace("/", "").replace("&", "").replace(",", "").isupper() and len(row[1]) > 5 and row[1] != "#"):
        category_rows.add(i)

# Columns we care about (0-indexed in CSV): 1=#, 2=Category, 3=Functionality, 4=Description, 5=Third-Party, 6=Notes, 8=Standard, 9=Extra, 10=Comments
# Write to XLSX starting from row 1
col_map = [1, 2, 3, 4, 5, 6, 8, 9, 10]  # CSV column indices
col_headers = ["#", "Category", "Functionality", "Description", "Third-Party Integration", "Notes", "Provided as Standard (Yes or No)", "(Can be) Provided as Extra (Yes or No)", "Comments"]
col_widths = [5, 22, 25, 50, 20, 35, 14, 14, 60]

# Write header row
for c_idx, hdr in enumerate(col_headers, 1):
    cell = ws.cell(row=1, column=c_idx, value=hdr)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    cell.border = thin_border

# Set column widths
for c_idx, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(c_idx)].width = w

# Resolve [VERIFY] tags
def resolve_verify(text):
    """Replace [VERIFY] tags with honest notes."""
    if not text:
        return text
    text = text.replace("[VERIFY audit firm and last audit date]", "Audit firm and last audit date to be confirmed during implementation planning.")
    text = text.replace("[VERIFY]", "To be confirmed during implementation planning.")
    text = text.replace("[VERIFY JFSC format support]", "JFSC statutory format support to be confirmed during implementation planning.")
    text = text.replace("[VERIFY SSE availability in current release]", "SSE availability to be confirmed against current release.")
    text = text.replace("[VERIFY contractual testing cadence]", "Testing cadence defined per deployment agreement.")
    # Replace em dashes with commas or colons
    text = text.replace(" — ", ", ")
    text = text.replace("—", ", ")
    return text

# Write data rows
xlsx_row = 2
for i, row in enumerate(rows):
    if i == 0:  # Skip first empty row
        continue
    if header_idx is not None and i == header_idx:
        continue  # Skip header row (already written)
    if len(row) < 2:
        continue

    # Check if this is a category separator row
    val1 = row[1].strip() if len(row) > 1 else ""
    is_category = (val1 and val1.isupper() and len(val1) > 3 and val1 != "#" and not val1.isdigit())
    
    # Skip the "Platform Functionalities" title row and empty rows
    if val1 == "" and all(c.strip() == "" for c in row):
        continue
    if val1 == "Platform Functionalities":
        continue
    if val1 == "" and len(row) > 2 and row[2].strip() == "":
        continue

    if is_category:
        # Write category separator row spanning all columns
        cell = ws.cell(row=xlsx_row, column=1, value=val1)
        for c in range(1, len(col_headers) + 1):
            cell = ws.cell(row=xlsx_row, column=c)
            if c == 1:
                cell.value = val1
            cell.font = cat_font
            cell.fill = cat_fill
            cell.border = thin_border
        ws.merge_cells(start_row=xlsx_row, start_column=1, end_row=xlsx_row, end_column=len(col_headers))
        xlsx_row += 1
        continue

    # Regular data row
    for c_idx, csv_col in enumerate(col_map):
        if csv_col < len(row):
            val = resolve_verify(row[csv_col].strip())
        else:
            val = ""
        
        cell = ws.cell(row=xlsx_row, column=c_idx + 1, value=val)
        cell.font = data_font
        cell.alignment = wrap_align
        cell.border = thin_border
        
        # Color coding for Standard/Extra columns
        col_header = col_headers[c_idx]
        if "Standard" in col_header or "Extra" in col_header:
            cell.alignment = Alignment(horizontal="center", vertical="top")
            if val == "Yes":
                cell.font = yes_font
                cell.fill = yes_fill
            elif val == "No":
                cell.font = no_font
                cell.fill = no_fill
            elif val == "Partial":
                cell.font = partial_font
                cell.fill = partial_fill

    xlsx_row += 1

# Freeze header row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(col_headers))}{xlsx_row - 1}"

wb.save(OUTPUT_XLSX)
print(f"Saved to {OUTPUT_XLSX}")
print(f"Total data rows: {xlsx_row - 2}")
