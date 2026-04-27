#!/usr/bin/env python3
"""Run official-style questionnaire row-fill regression for bid-manager."""
from __future__ import annotations

import argparse
import json
import shutil
import subprocess
import time
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parents[2]
BID = ROOT / "legacy" / "bid-manager"

HEADERS = [
    "Ref #",
    "Category",
    "Sub-Category",
    "Question / Requirement",
    "SettleMint Response",
    "Status",
    "Evidence / Reference",
    "Buyer-only remarks",
]

ROWS = [
    ["Q1", "Platform", "Tokenization", "Describe support for EVM tokenized asset lifecycle workflows.", "", "", "", "Do not edit"],
    ["Q2", "Security", "Governance", "Explain maker-checker controls and audit evidence.", "", "", "", "Do not edit"],
    ["Q3", "Integration", "Custody", "Describe custody/signing integration posture and operating boundary.", "", "", "", "Do not edit"],
]

RESPONSES = {
    "Q1": {
        "response": "DALP supports EVM-based tokenized asset lifecycle workflows, including issuance, transfer controls, servicing events, and role-based operational oversight.",
        "status": "Met",
        "evidence": "DALP canon; bid-manager regression fixture",
    },
    "Q2": {
        "response": "The operating model supports maker-checker governance, permissioned actions, audit trails, and evidence suitable for internal risk and control review.",
        "status": "Met",
        "evidence": "Bid-manager technical skeleton; governance response pattern",
    },
    "Q3": {
        "response": "Custody and signing are handled through the bank-approved custodian or signing provider boundary; DALP integrates into the controlled operating model rather than claiming to be the legal custodian.",
        "status": "Met",
        "evidence": "Bid-manager custody boundary rule",
    },
}


def run_cmd(name: str, cmd: list[str]) -> dict[str, Any]:
    start = time.time()
    proc = subprocess.run(cmd, text=True, capture_output=True)
    return {
        "name": name,
        "cmd": cmd,
        "exit": proc.returncode,
        "ok": proc.returncode == 0,
        "duration_seconds": round(time.time() - start, 3),
        "stdout_tail": proc.stdout[-4000:],
        "stderr_tail": proc.stderr[-4000:],
    }


def create_source(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for row in ROWS:
        ws.append(row)
    dv = DataValidation(type="list", formula1='"Met,Partially Met,Not Met,Clarification"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"F2:F{len(ROWS)+1}")
    ws.protection.sheet = True
    for col, width in {"A": 12, "B": 18, "C": 18, "D": 58, "E": 70, "F": 18, "G": 36, "H": 24}.items():
        ws.column_dimensions[col].width = width
    wb.save(path)


def fill_workbook(source: Path, final: Path) -> dict[str, Any]:
    shutil.copy2(source, final)
    wb = load_workbook(final)
    ws = wb["Requirements"]
    header = [cell.value for cell in ws[1]]
    positions = {name: idx + 1 for idx, name in enumerate(header)}
    required = {"Ref #", "SettleMint Response", "Status", "Evidence / Reference", "Buyer-only remarks"}
    missing = sorted(required - set(positions))
    if missing:
        raise RuntimeError(f"Missing expected columns: {missing}")
    for row in range(2, ws.max_row + 1):
        ref = ws.cell(row=row, column=positions["Ref #"]).value
        if ref in RESPONSES:
            answer = RESPONSES[ref]
            ws.cell(row=row, column=positions["SettleMint Response"]).value = answer["response"]
            ws.cell(row=row, column=positions["Status"]).value = answer["status"]
            ws.cell(row=row, column=positions["Evidence / Reference"]).value = answer["evidence"]
    wb.save(final)
    return verify_workbook(source, final)


def zip_parts(path: Path) -> set[str]:
    with zipfile.ZipFile(path) as zf:
        bad = zf.testzip()
        if bad:
            raise RuntimeError(f"Bad zip entry {bad}")
        return set(zf.namelist())


def verify_workbook(source: Path, final: Path) -> dict[str, Any]:
    source_parts = zip_parts(source)
    final_parts = zip_parts(final)
    wb = load_workbook(final, data_only=False)
    ws = wb["Requirements"]
    header = [cell.value for cell in ws[1]]
    positions = {name: idx + 1 for idx, name in enumerate(header)}
    response_count = status_count = evidence_count = buyer_remarks_untouched = 0
    statuses = []
    for row in range(2, ws.max_row + 1):
        ref = ws.cell(row=row, column=positions["Ref #"]).value
        response = ws.cell(row=row, column=positions["SettleMint Response"]).value
        status = ws.cell(row=row, column=positions["Status"]).value
        evidence = ws.cell(row=row, column=positions["Evidence / Reference"]).value
        remark = ws.cell(row=row, column=positions["Buyer-only remarks"]).value
        if response:
            response_count += 1
        if status:
            status_count += 1
            statuses.append(status)
        if evidence:
            evidence_count += 1
        if remark == "Do not edit":
            buyer_remarks_untouched += 1
    return {
        "source_parts": len(source_parts),
        "final_parts": len(final_parts),
        "parts_preserved": source_parts == final_parts,
        "rows": len(ROWS),
        "response_count": response_count,
        "status_count": status_count,
        "evidence_count": evidence_count,
        "statuses": statuses,
        "buyer_remarks_untouched": buyer_remarks_untouched,
        "openpyxl_load_ok": True,
    }


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Run bid-manager official-style questionnaire row-fill regression")
    parser.add_argument("--json", action="store_true")
    parser.add_argument("--out-dir", default=None)
    ns = parser.parse_args(argv)

    stamp = time.strftime("%Y%m%dT%H%M%SZ", time.gmtime())
    out_dir = Path(ns.out_dir).resolve() if ns.out_dir else BID / "output" / f"questionnaire-regression-{stamp}"
    out_dir.mkdir(parents=True, exist_ok=True)
    source = out_dir / "bid-manager-official-questionnaire-source.xlsx"
    final = out_dir / "bid-manager-official-questionnaire-filled.xlsx"
    create_source(source)
    verification = fill_workbook(source, final)
    render = run_cmd("libreoffice_xlsx_to_pdf", ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", str(out_dir), str(final)])
    pdf = out_dir / (final.stem + ".pdf")
    verification["libreoffice_render_ok"] = render["ok"] and pdf.exists()
    verification["render_pdf"] = str(pdf)
    ok = bool(
        verification["parts_preserved"]
        and verification["response_count"] == len(ROWS)
        and verification["status_count"] == len(ROWS)
        and verification["evidence_count"] == len(ROWS)
        and verification["buyer_remarks_untouched"] == len(ROWS)
        and verification["libreoffice_render_ok"]
    )
    payload = {
        "agent": "bid-manager",
        "artifact_dir": str(out_dir),
        "source": str(source),
        "final": str(final),
        "render": render,
        "verification": verification,
        "ok": ok,
    }
    (out_dir / "questionnaire-regression.json").write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(json.dumps(payload, indent=2) if ns.json else f"bid-manager questionnaire regression: {'OK' if ok else 'FAIL'} artifacts={out_dir}")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
