#!/usr/bin/env python3
"""Convert questionnaire CSV to the approved locked XLSX template.

This path is intentionally fail-closed:
- never creates questionnaire workbooks from scratch
- always clones the approved workbook template
- only populates the Response sheet data rows
- validates the final workbook against the locked template contract

Usage:
    python3 scripts/csv_to_xlsx.py input.csv [output.xlsx]
"""

from __future__ import annotations

import sys
from pathlib import Path
import shutil

from openpyxl import load_workbook

SHARED_SCRIPTS_DIR = Path(__file__).resolve().parents[2] / "shared" / "scripts"
if str(SHARED_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SHARED_SCRIPTS_DIR))

from office_output_validation import (
    OfficeValidationError,
    read_csv_rows,
    validate_csv_has_header,
    validate_questionnaire_template,
    validate_xlsx_expected_shape,
)
from office_runtime_guard import append_build_metadata, assert_output_within_agent


TEMPLATE_PATH = (
    Path(__file__).resolve().parents[1]
    / "skeletons"
    / "5_questionnaire"
    / "excel"
    / "settlemint-questionnaire-template.xlsx"
)
RESPONSE_SHEET_NAME = "Response"
RESPONSE_HEADER_ROW = 2
RESPONSE_DATA_START_ROW = 3


def _normalized_row(values: list[object]) -> list[str]:
    normalized: list[str] = []
    for value in values:
        if value is None:
            normalized.append("")
        else:
            normalized.append(str(value).strip())
    return normalized


def _clear_response_rows(ws) -> None:
    for row in range(RESPONSE_DATA_START_ROW, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).value = None


def convert(input_path: str, output_path: str | None = None) -> Path:
    src = Path(input_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")

    dst = Path(output_path) if output_path else src.with_suffix(".xlsx")
    dst = assert_output_within_agent("bid-manager", dst)
    dst.parent.mkdir(parents=True, exist_ok=True)

    rows = read_csv_rows(src)
    validate_csv_has_header(rows, src)

    template_path = validate_questionnaire_template(TEMPLATE_PATH)
    template_wb = load_workbook(template_path)
    response_ws = template_wb[RESPONSE_SHEET_NAME]
    template_headers = _normalized_row(
        [response_ws.cell(row=RESPONSE_HEADER_ROW, column=col).value for col in range(1, response_ws.max_column + 1)]
    )

    csv_headers = _normalized_row(rows[0])
    if csv_headers != template_headers:
        raise OfficeValidationError(
            "Questionnaire CSV header/order must match the approved template exactly. "
            f"Expected {template_headers}, got {csv_headers}"
        )

    response_capacity = response_ws.max_row - RESPONSE_DATA_START_ROW + 1
    data_rows = rows[1:]
    if len(data_rows) > response_capacity:
        raise OfficeValidationError(
            "Questionnaire exceeds the approved workbook capacity. "
            f"Template allows {response_capacity} response rows, got {len(data_rows)}"
        )

    shutil.copy2(template_path, dst)
    wb = load_workbook(dst)
    ws = wb[RESPONSE_SHEET_NAME]
    _clear_response_rows(ws)

    for row_idx, row in enumerate(data_rows, start=RESPONSE_DATA_START_ROW):
        padded_row = list(row) + [""] * (ws.max_column - len(row))
        for col_idx in range(1, ws.max_column + 1):
            value = padded_row[col_idx - 1]
            ws.cell(row=row_idx, column=col_idx).value = value if value != "" else None

    # Apply SettleMint branding to data rows
    from excel_brand import apply_brand_to_existing, auto_column_widths
    apply_brand_to_existing(
        ws,
        header_row=RESPONSE_HEADER_ROW,
        data_start_row=RESPONSE_DATA_START_ROW,
        data_end_row=RESPONSE_DATA_START_ROW + len(data_rows) - 1,
    )
    auto_column_widths(ws, col_start=1, col_end=ws.max_column)

    wb.save(dst)
    validate_xlsx_expected_shape(
        dst,
        template_path=template_path,
        expected_response_headers=template_headers,
        expected_response_rows=len(data_rows),
    )
    append_build_metadata(
        "bid-manager",
        output_path=dst,
        operation="csv_to_xlsx",
        extra={
            "input": str(src),
            "template": str(template_path),
            "response_rows": len(data_rows),
        },
    )
    print(
        f"✓ Converted {src.name} → {dst.name} "
        f"({len(data_rows)} response rows, approved template cloned, workbook contract verified)"
    )
    return dst


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__.strip())
        sys.exit(1)
    convert(sys.argv[1], sys.argv[2] if len(sys.argv) > 2 else None)
