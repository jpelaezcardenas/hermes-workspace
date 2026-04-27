#!/usr/bin/env python3
"""
Fill in columns I (Provided as Standard), J (Can be Provided as Extra), K (Comments)
in the Neptune XLSX questionnaire for DALP.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import shutil
import os

INPUT_PATH = "/Users/quark/.openclaw/media/inbound/4c25ab05-5ce1-4709-9a7f-37d200a415d0.xlsx"
OUTPUT_PATH = "/Users/quark/Public/quark/workspace/quark-unsorted-output/Neptune_Questionnaire_Response_DALP.xlsx"

# Ensure output dir exists
os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)

# Copy input to output (preserve all original content and formatting)
shutil.copy2(INPUT_PATH, OUTPUT_PATH)

wb = openpyxl.load_workbook(OUTPUT_PATH)
ws = wb["Platform Functionalities"]

# Build answers: row_number -> (col_I, col_J, col_K)
# Row numbers correspond to actual Excel rows. We'll detect them by looking at column B values.
# Format: question_number -> (standard, extra, comment)
# standard/extra: "Yes" / "No" / "Partial"

ANSWERS = {
    # TOKEN ISSUANCE & LIFECYCLE
    1: ("Yes", "No",
        "DALP deploys ERC-3643 compliant smart contracts on any EVM-compatible chain via the Asset Designer UI; token name, symbol, decimals, supply cap, and standard are all configurable at creation time."),
    2: ("Yes", "No",
        "DALP enforces role-gated minting: only wallets with the mint role can execute mint operations, and recipient wallets must hold verified identity claims before tokens can be issued."),
    3: ("Yes", "No",
        "DALP implements on-chain token burning via the redeem or burn function; total supply decreases are recorded on-chain and reflected in the cap table immediately."),
    4: ("Yes", "No",
        "DALP deploys a global pause function on every token contract; activation requires the emergency role, providing a circuit-breaker that halts all transfers platform-wide."),
    5: ("Yes", "No",
        "DALP supports per-address freeze (full or partial) via on-chain freeze functions enforced by the compliance layer before every transfer; granular control is available through the compliance officer role."),
    6: ("Yes", "No",
        "DALP implements forced transfer as a role-gated on-chain operation with a complete immutable audit trail; every forced transfer is logged with actor identity, timestamp, and reason."),
    7: ("Yes", "No",
        "DALP provides a durable identity-recovery and token-recovery workflow: the platform orchestrates wallet redeployment, identity migration, and token reissuance after multi-step identity re-verification."),
    8: ("Yes", "No",
        "DALP tracks total and circulating supply in real time via on-chain event indexing; supply data including caps, reserves, and treasury allocations is displayed in the issuer dashboard."),
    9: ("Yes", "No",
        "DALP supports concurrent deployment of equity, bond, fund, stablecoin, deposit, real estate, and precious metal token types within the same platform instance, each with class-specific validation."),
    10: ("Yes", "No",
        "DALP provides a per-token document vault for storing legal documents, prospectus links, and terms; every uploaded document is SHA-256 hashed for integrity verification, and access is visibility-controlled."),
    # COMPLIANCE & IDENTITY
    11: ("Partial", "Yes",
        "DALP provides a built-in KYC review workflow and claim issuance engine; integration with third-party KYC/AML providers such as Sumsub or Onfido is configured by the platform operator to automate identity verification."),
    12: ("Yes", "No",
        "DALP uses ONCHAINID (ERC-3643 standard) to anchor each investor's verified identity on-chain; validated claims including KYC status, accreditation, and jurisdiction are stored against each investor's on-chain identity contract."),
    13: ("Yes", "No",
        "DALP deploys a per-system Identity Registry smart contract mapping wallet addresses to verified ONCHAINID contracts; every token transfer checks this registry to confirm recipient eligibility."),
    14: ("Yes", "No",
        "DALP maintains a Trusted Issuers Registry as an on-chain smart contract; platform administrators configure which entities (KYC providers, regulators) are authorized to issue identity claims, with a global registry tier for platform-wide trusted issuers."),
    15: ("Yes", "No",
        "DALP maintains a configurable Claim Topics Registry; operators define the required claim types per token, and the compliance engine enforces them before every transfer using a deterministic module evaluation chain."),
    16: ("Yes", "No",
        "DALP implements 12 compliance module types as modular smart contracts including jurisdiction controls, address restrictions, investor caps, time-based rules, and supply controls; all rules are checked before every transfer."),
    17: ("Partial", "Yes",
        "DALP does not perform autonomous sanctions list lookups natively; sanctions screening is delivered via integration with providers such as Sumsub or Chainalysis, with screening results recorded as on-chain identity claims that block transfers automatically."),
    18: ("Partial", "Yes",
        "DALP's claim lifecycle supports recording accredited or qualified investor status as an on-chain claim with expiry; verification workflows and renewal checks are delivered via integration with providers such as Sumsub or VerifyInvestor."),
    19: ("Yes", "No",
        "DALP's CountryRestriction compliance module enforces country-level whitelisting and blacklisting at the smart contract level; any transfer involving a restricted jurisdiction is blocked automatically before execution."),
    20: ("Yes", "No",
        "DALP implements a MaxInvestors compliance module that enforces a hard cap on the number of eligible token holders per token; transfers that would exceed the configured cap are rejected on-chain."),
    21: ("No", "Yes",
        "DALP does not perform PEP or adverse media screening natively; this is delivered via integration with providers such as Sumsub or Refinitiv, with screening results recorded as on-chain identity claims consumed by the compliance layer."),
    # CORPORATE ACTIONS
    22: ("Partial", "Yes",
        "DALP's fixed-treasury-yield feature supports scheduled ERC-20 token distributions (including USDC) from a treasury to all eligible token holders; the distribution schedule, amounts, and record date are configurable per token issuance."),
    23: ("Partial", "Yes",
        "DALP's fixed-treasury-yield feature supports scheduled fixed-rate coupon payments for bond tokens; the yield schedule, payment frequency, and denomination currency are configured at token creation or updated by an authorized operator."),
    24: ("Partial", "Yes",
        "DALP implements ERC-5805 voting power infrastructure with checkpointed token balances and on-chain delegation; a full governance proposal-and-execution system is not yet shipped, but the voting power foundation enables integration with governance frameworks."),
    25: ("Yes", "No",
        "DALP's ERC-5805 voting power feature provides on-chain checkpointed snapshots of all holder balances at any historical block; these immutable snapshots underpin entitlement calculations for dividends, voting, and rights issues."),
    26: ("No", "Yes",
        "DALP does not include a native capital call workflow for fund or SPC structures; capital calls can be approximated using the token-sale addon to issue additional tokens to a whitelist of existing investors. [VERIFY]"),
    27: ("Yes", "No",
        "DALP's maturity-redemption feature burns tokens and triggers treasury payouts to holders at maturity; issuer buybacks can be executed via forced transfer followed by burning with a full on-chain audit trail."),
    28: ("No", "No",
        "DALP does not provide a native stock split or token consolidation function; proportional supply reorganization at the contract level is not currently supported as a built-in operation."),
    29: ("Partial", "Yes",
        "DALP's token-sale addon supports presale whitelisting of existing holders with a configurable discount and per-address cap, enabling a rights-issue-style offering to current investors before a public sale."),
    30: ("Partial", "Yes",
        "DALP supports on-chain document storage and metadata anchoring via the token document vault for publishing announcements and material notices; a dedicated on-chain broadcast function with investor notification is not natively provided."),
    31: ("No", "Yes",
        "DALP does not calculate or withhold taxes natively; distributions are paid gross and tax reporting responsibility lies with the platform operator or their integrated tax service provider."),
    # INVESTOR ONBOARDING & MGMT
    32: ("Yes", "No",
        "DALP provides a white-label investor portal with self-service registration, identity document submission, and guided KYC onboarding; the interface supports per-tenant branding including custom domain, logo, and color scheme."),
    33: ("Partial", "Yes",
        "DALP's token-sale addon delivers a digital subscription workflow with KYC verification gating and ERC-20 payment processing; e-signature integration with DocuSign or Adobe Sign for subscription agreement signing can be added as an extra."),
    34: ("Yes", "No",
        "DALP provides each investor with a personalized dashboard displaying portfolio holdings, transaction history, and distribution records derived from real-time on-chain state."),
    35: ("Yes", "No",
        "DALP supports self-custodied wallets via WalletConnect (MetaMask, Ledger), platform-managed wallets, and institutional custody via Fireblocks; investors select the wallet type appropriate to their profile."),
    36: ("Partial", "Yes",
        "DALP generates in-platform notifications for key events; external email delivery and an auditable communication log require configuration of the operator's email notification service."),
    37: ("Yes", "No",
        "DALP provides a per-token document vault with visibility-controlled access; investors can access their disclosure documents, and all uploads are SHA-256 hashed and versioned."),
    38: ("Partial", "Yes",
        "DALP's compliance module system supports defining different eligibility criteria and allocation limits per investor category; a dedicated tiering UI with named JFSC categories and per-tier fee structures requires additional operator configuration."),
    # CUSTODY & WALLET INFRASTRUCTURE
    39: ("Yes", "No",
        "DALP integrates natively with Fireblocks (MPC-CMP) and DFNS (threshold MPC) as institutional custody providers; transaction signing, policy enforcement, and key management are delegated to the configured provider."),
    40: ("Yes", "No",
        "DALP's vault addon supports configurable multi-signature wallets with configurable quorum thresholds; Fireblocks multi-approval and DFNS policy-engine approval workflows are also supported."),
    41: ("Yes", "No",
        "DALP enforces transaction authorization at two layers: DALP-native TransferPolicy (amount limits, wallet verification gates) and the custody provider's own TAP (Fireblocks) or programmatic policy engine (DFNS)."),
    42: ("Partial", "Yes",
        "DALP supports segregated wallet configurations; the hot/warm/cold operational strategy is implemented through the custody provider's vault account hierarchy (Fireblocks), which is configurable by the operator."),
    43: ("Yes", "No",
        "DALP enforces wallet whitelisting at the smart contract level via the identity registry; only wallets with a verified on-chain identity and required claims can receive tokens, enforced on every transfer."),
    44: ("Yes", "No",
        "DALP provides a governed identity-recovery and token-recovery workflow for lost or compromised wallets; key reconstruction for MPC wallets is managed by the integrated custody provider (Fireblocks or DFNS)."),
    # SECONDARY MARKET & TRADING
    45: ("No", "Yes",
        "DALP does not include a native peer-to-peer order posting or OTC marketplace; secondary trading facilitation is outside the current platform scope. Integration with an external regulated ATS or marketplace is required."),
    46: ("Yes", "No",
        "DALP's XvP (Exchange versus Payment) Settlement addon provides atomic DvP and PvP settlement with escrow-based approval and all-or-nothing execution; counterparty risk is eliminated by design."),
    47: ("Yes", "No",
        "DALP generates on-chain settlement records for every XvP execution; all settlement events are indexed in real time and exportable via the REST API for compliance and accounting system feeds."),
    48: ("Yes", "No",
        "DALP enforces time-based transfer restrictions via the TimeTransfer compliance module at the smart contract level; lock-up periods expire automatically on the configured date without manual intervention."),
    49: ("No", "No",
        "DALP does not provide an order book or automated trade matching engine; DALP is designed as an asset issuance and lifecycle management platform, not a trading venue. An external ATS or MTF is required for order matching."),
    50: ("No", "Yes",
        "DALP does not include pre-built connectors to regulated secondary exchanges; integration with venues such as 21X or Archax can be built by operators using DALP's comprehensive REST API, but is not a packaged integration."),
    # CAP TABLE & REPORTING
    51: ("Yes", "No",
        "DALP maintains a live cap table derived from on-chain token holder balances; the cap table reflects wallet addresses, verified identities, jurisdictions, holdings, and percentage ownership in real time."),
    52: ("Yes", "No",
        "DALP provides a complete, immutable on-chain audit trail of all token events including transfers, minting, burning, corporate actions, compliance decisions, and role changes; all records are indexed and exportable."),
    53: ("Partial", "Yes",
        "DALP provides manual CSV and JSON data exports plus a compliance dashboard for operator-driven reporting; automated regulatory report generation and jurisdiction-specific filing templates (JFSC) are not yet natively implemented."),
    54: ("Partial", "Yes",
        "DALP enables investors to view holdings and distributions in the investor dashboard; automated periodic report generation (PDF statements, tax certificates) is not yet natively shipped and requires operator-side tooling."),
    55: ("Yes", "No",
        "DALP provides a real-time compliance dashboard for compliance officers showing KYC status, pending verifications, sanctions alerts, and transfer restriction events with drill-down to individual investor records."),
    56: ("Partial", "Yes",
        "DALP exposes a comprehensive REST API and CSV exports for financial data; pre-built connectors to specific ERP or accounting systems are not included and must be implemented by the operator or integration partner."),
    # PLATFORM ADMIN & SECURITY
    57: ("Yes", "No",
        "DALP implements a multi-layer RBAC model combining on-chain smart contract roles (governance, systemManager, identityManager, and others) with off-chain API permissions; least-privilege is enforced at every API endpoint and contract function."),
    58: ("Yes", "No",
        "DALP requires multi-factor wallet verification (TOTP, PIN codes, or secret codes) for all privileged transaction signing; passkey (WebAuthn) authentication is available for platform login."),
    59: ("Partial", "Yes",
        "DALP supports OIDC-based authentication via Better Auth; enterprise SAML 2.0 SSO integration with providers such as Okta or Azure AD is available as a configurable extra through the Better Auth plugin system."),
    60: ("Yes", "No",
        "DALP logs every user action, API call, and on-chain event with timestamp, user identity, and action details; logs are tamper-evident and retained in the database for the configured retention period."),
    61: ("Yes", "No",
        "DALP maintains a formal audit slice system tracking all smart contract surfaces in scope for independent security review; contracts are audited before mainnet deployment and re-audited after material code changes. [VERIFY audit firm and last audit date]"),
    62: ("Yes", "No",
        "DALP deployments are protected by Cloudflare or AWS WAF for DDoS mitigation, rate limiting, and web application firewall protection; anomaly detection and intrusion protection are standard in the SettleMint-managed deployment topology."),
    63: ("Yes", "No",
        "DALP undergoes regular independent penetration testing; critical and high findings are remediated before go-live. Testing cadence and scope are defined in the deployment and maintenance agreement. [VERIFY contractual testing cadence]"),
    64: ("Yes", "No",
        "DALP encrypts all data in transit via TLS and at rest via platform-managed secret backends (AWS Secrets Manager, GCP Secret Manager, Azure Key Vault, or HashiCorp Vault); field-level PII encryption is supported."),
    65: ("Partial", "Yes",
        "DALP supports multi-region cloud deployment and automated failover configurations; specific RPO and RTO SLA commitments (such as RPO below 1 hour and RTO below 4 hours) are defined and contractually documented per managed deployment agreement."),
    66: ("Yes", "No",
        "DALP provides a secure REST API layer with API key authentication, HTTP-method-based scope enforcement, read-only versus read-write key differentiation, and configurable per-key rate limiting."),
    # INTEGRATION & INTEROPERABILITY
    67: ("Yes", "No",
        "DALP operates on any EVM-compatible chain including Ethereum, Polygon, Avalanche C-Chain, Arbitrum, Optimism, Base, BNB Smart Chain, and Hyperledger Besu; multi-chain deployment is supported with per-chain identity, compliance, and indexer isolation. Non-EVM chains are not supported."),
    68: ("Yes", "No",
        "DALP integrates with ERC-20 stablecoins including USDC for on-chain distribution payments, subscription funding, and trade settlement; payment currency is configurable per token or primary offering."),
    69: ("No", "Yes",
        "DALP does not include native fiat banking rails; fiat on/off-ramp requires integration with a licensed third-party payment processor or banking API configured by the platform operator."),
    70: ("Partial", "Yes",
        "DALP includes an exchange rate feeds system for ingesting external FX data; integration with on-chain oracle networks such as Chainlink or API3 for NAV pricing and real-world asset valuation can be configured as an extra."),
    71: ("Yes", "No",
        "DALP provides a comprehensive REST API (v2) with full OpenAPI contract coverage, a typed SDK, and an oRPC endpoint; the API covers all platform functions including token operations, compliance, identity, reporting, and admin."),
    72: ("Partial", "Yes",
        "DALP provides Server-Sent Events (SSE) streaming for real-time monitoring data; a full outbound webhook delivery system for pushing events to external systems on demand is not yet natively implemented. [VERIFY SSE availability in current release]"),
    73: ("No", "Yes",
        "DALP does not include a pre-built ERP or accounting system connector; financial data is accessible via the REST API and CSV exports, and a custom integration with accounting or leasing management systems can be built by the operator."),
    # MULTI-TENANCY
    74: ("Yes", "No",
        "DALP supports admin-assisted tenant onboarding via the organization management UI; each tenant receives an isolated environment with configurable settings including blockchain selection, token standards, base currency, and theme."),
    75: ("Yes", "No",
        "DALP supports per-organization theme configuration including custom logo, color scheme, and CSS variables; custom domain configuration is supported, and SSL certificates can be provisioned per tenant deployment."),
    76: ("Yes", "No",
        "DALP enforces data isolation between tenants via organizationId scoping on all database queries and through per-tenant on-chain system contract deployments; cross-tenant data access is architecturally prevented at both the application and contract layers."),
    77: ("Yes", "No",
        "DALP separates platform-admin roles from per-tenant owner and member roles; each tenant independently manages their own users and role assignments, and platform admins cannot access tenant-level data without an explicit audit-logged authorization."),
    78: ("Partial", "Yes",
        "DALP tracks per-organization API usage and on-chain asset counts; a configurable billing engine with fee schedules, automated invoice generation, and tenant revenue reporting is not natively included and requires operator-side integration."),
    79: ("Yes", "No",
        "DALP's compliance module system is fully tenant-scoped; each organization independently configures KYC/AML provider selection, required claim types, permitted investor jurisdictions, and investor eligibility criteria with no cross-tenant interference."),
    80: ("Partial", "Yes",
        "DALP provides a platform-admin organization listing view with on-chain asset counts and member metrics per tenant; a consolidated revenue dashboard with billing metrics and full tenant lifecycle management (suspend/offboard) is not yet natively shipped."),
    # SPV / LEGAL ENTITY MANAGEMENT
    81: ("No", "Yes",
        "DALP does not include a native SPV legal entity registry with structured fields for incorporation details, directors, and registered agents; token issuances are managed in DALP while SPV legal records would need to be maintained externally or via a custom operator configuration. [VERIFY]"),
    82: ("Partial", "Yes",
        "DALP supports associating metadata (classification, location, and identifier claims) with each token issuance; maritime-specific fields such as IMO number and vessel class can be stored as custom token metadata or in the token document vault rather than in structured database fields."),
    83: ("Partial", "Yes",
        "DALP's token document vault stores and versions issuance-level legal documents (offering memorandum, legal opinions, shareholder agreements) with visibility-controlled access for ISSUER, COMPLIANCE, and AUDITOR roles; a pre-configured SPV-specific document taxonomy is not natively provided."),
    84: ("No", "Yes",
        "DALP does not provide a visual ownership chain diagram natively; the platform records token-level data but does not render a multi-layer SPV-to-asset structure visualization for admin or investor-facing views. [VERIFY]"),
    85: ("Partial", "Yes",
        "DALP supports the core lifecycle events of a token issuance including active, paused, maturity redemption, and token burning on wind-down; formal mapping to SPV legal lifecycle stages and triggering of associated legal actions requires operator configuration."),
    86: ("Partial", "Yes",
        "DALP's real-time cap table records all token holders with identity, holdings, percentage ownership, and acquisition data from on-chain state; automated generation and export of a JFSC-compliant beneficial ownership register in statutory filing format requires operator-side tooling or a custom integration. [VERIFY JFSC format support]"),
}

print(f"Loaded {len(ANSWERS)} answers.")

# Now find the rows with numbered items and fill in columns I, J, K
filled = 0
for row in ws.iter_rows():
    # Column B is index 2 (1-based)
    b_cell = row[1]  # B column
    b_val = b_cell.value
    
    # Check if col B has a number (1-86)
    if b_val is not None:
        try:
            q_num = int(str(b_val).strip())
            if 1 <= q_num <= 86 and q_num in ANSWERS:
                row_num = b_cell.row
                std, extra, comment = ANSWERS[q_num]
                
                # Col I = column 9, J = 10, K = 11
                ws.cell(row=row_num, column=9, value=std)
                ws.cell(row=row_num, column=10, value=extra)
                ws.cell(row=row_num, column=11, value=comment)
                
                # Apply wrapping to comment cell
                ws.cell(row=row_num, column=11).alignment = Alignment(wrap_text=True, vertical='top')
                
                filled += 1
        except (ValueError, TypeError):
            pass

print(f"Filled {filled} rows.")

wb.save(OUTPUT_PATH)
print(f"Saved to {OUTPUT_PATH}")
