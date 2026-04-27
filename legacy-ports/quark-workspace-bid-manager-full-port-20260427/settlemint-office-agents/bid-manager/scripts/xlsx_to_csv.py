#!/usr/bin/env python3
"""Convert XLSX to CSV.

Primary method: Microsoft's markitdown (outputs markdown tables).
Fallback: openpyxl direct extraction (produces standard CSV).

For bid-manager questionnaire processing, the openpyxl path is preferred
because it produces clean CSV that's easier to parse programmatically.
markitdown is available as an alternative for richer formatting context.

Usage:
    python3 scripts/xlsx_to_csv.py input.xlsx [output.csv]
    python3 scripts/xlsx_to_csv.py input.xlsx [output.csv] --markitdown
"""

import csv
import io
import subprocess
import sys
from pathlib import Path

MARKITDOWN = "/Users/quark/Library/Python/3.14/bin/markitdown"


def _convert_openpyxl(src: Path, dst: Path) -> None:
    """Convert using openpyxl — clean CSV output, best for structured data."""
    from openpyxl import load_workbook

    wb = load_workbook(str(src), read_only=True, data_only=True)
    buf = io.StringIO()
    writer = csv.writer(buf)

    for idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        if len(wb.sheetnames) > 1:
            if idx > 0:
                writer.writerow([])  # blank line between sheets
            writer.writerow([f"## Sheet: {sheet_name}"])

        for row in ws.iter_rows(values_only=True):
            writer.writerow([(cell if cell is not None else "") for cell in row])

    wb.close()
    dst.write_text(buf.getvalue(), encoding="utf-8")


def _convert_markitdown(src: Path, dst: Path) -> None:
    """Convert using markitdown — markdown table output with formatting context."""
    result = subprocess.run(
        [MARKITDOWN, str(src)],
        capture_output=True,
        text=True,
        timeout=120,
    )
    if result.returncode != 0:
        raise RuntimeError(
            f"markitdown failed (exit {result.returncode}): {result.stderr.strip()}"
        )
    # markitdown outputs markdown; write as-is (caller gets markdown tables)
    dst.write_text(result.stdout, encoding="utf-8")


def convert(
    input_path: str,
    output_path: str | None = None,
    use_markitdown: bool = False,
) -> Path:
    src = Path(input_path)
    if not src.exists():
        raise FileNotFoundError(f"Input file not found: {src}")
    if src.suffix.lower() != ".xlsx":
        raise ValueError(f"Expected .xlsx file, got: {src.suffix}")

    default_ext = ".md" if use_markitdown else ".csv"
    dst = Path(output_path) if output_path else src.with_suffix(default_ext)

    if use_markitdown:
        _convert_markitdown(src, dst)
        print(f"✓ Converted {src.name} → {dst.name} (markitdown)")
    else:
        _convert_openpyxl(src, dst)
        print(f"✓ Converted {src.name} → {dst.name} (openpyxl)")

    return dst


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__.strip())
        sys.exit(1)

    use_md = "--markitdown" in sys.argv
    args = [a for a in sys.argv[1:] if a != "--markitdown"]

    convert(args[0], args[1] if len(args) > 1 else None, use_markitdown=use_md)
